import requests
import json
from typing import Dict, List, Optional, Set
import time
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings
import numpy as np
from pybaseball import playerid_lookup, batting_stats, pitching_stats, statcast_batter, statcast_pitcher
from datetime import datetime, timedelta
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from collections import defaultdict
import re
import os
import sys
warnings.filterwarnings('ignore')

class OddsAPIExcelFetcher:
    def __init__(self, api_key: str):
        """
        Initialize the Odds API fetcher with Excel output capabilities
        
        Args:
            api_key (str): Your Odds API key
        """
        self.api_key = api_key
        self.base_url = "https://api.the-odds-api.com/v4"
        self.headers = {'Content-Type': 'application/json'}
        
        # The 12 prop categories you specified
        self.prop_categories = [
            'batter_home_runs',
            'batter_hits', 
            'batter_total_bases',
            'batter_rbis',
            'batter_runs_scored',
            'batter_hits_runs_rbis',
            'batter_singles',
            'batter_strikeouts',
            'pitcher_strikeouts',
            'pitcher_hits_allowed',
            'pitcher_earned_runs',
            'pitcher_outs'
        ]
        
        # Excel styling
        self.header_font = Font(bold=True, color='FFFFFF', size=11)
        self.header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        self.over_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # Light green for Over
        self.under_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')  # Light pink for Under
        self.data_font = Font(size=10)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Category colors for visual distinction
        self.category_colors = {
            'batter_home_runs': 'FFE699',
            'batter_hits': 'C5E0B4',
            'batter_total_bases': 'BDD7EE',
            'batter_rbis': 'F4B183',
            'batter_runs_scored': 'D5A6BD',
            'batter_hits_runs_rbis': 'A9D18E',
            'batter_singles': 'FFD966',
            'batter_strikeouts': 'F8CBAD',
            'pitcher_strikeouts': 'B4C7E7',
            'pitcher_hits_allowed': 'C9C9C9',
            'pitcher_earned_runs': 'F2CC8F',
            'pitcher_outs': 'E2EFDA'
        }
        
        # Initialize data structures for tracking
        self.unique_players = {}  # {player_name: {categories: [list], props: [list]}}
        self.over_props = {category: [] for category in self.prop_categories}
        self.under_props = {category: [] for category in self.prop_categories}
    
    def auto_adjust_columns(self, ws, max_width: int = 50):
        """
        Auto-adjust column widths, handling merged cells properly
        """
        for col_num in range(1, ws.max_column + 1):
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(col_num)
            
            for row_num in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_num, column=col_num)
                
                # Skip merged cells
                if hasattr(cell, 'coordinate') and cell.coordinate in ws.merged_cells:
                    continue
                
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            
            # Set column width
            adjusted_width = min(max_length + 2, max_width)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def get_mlb_events(self) -> List[Dict]:
        """
        Get all current MLB events/games
        """
        url = f"{self.base_url}/sports/baseball_mlb/events"
        params = {'apiKey': self.api_key}
        
        try:
            response = requests.get(url, params=params, headers=self.headers)
            response.raise_for_status()
            
            events = response.json()
            print(f"Found {len(events)} MLB events")
            return events
            
        except requests.exceptions.RequestException as e:
            print(f"Error fetching MLB events: {e}")
            return []
    
    def get_event_props(self, event_id: str, markets: List[str], 
                       regions: str = 'us', odds_format: str = 'american') -> Optional[Dict]:
        """
        Get prop odds for a specific event
        """
        url = f"{self.base_url}/sports/baseball_mlb/events/{event_id}/odds"
        params = {
            'apiKey': self.api_key,
            'regions': regions,
            'markets': ','.join(markets),
            'oddsFormat': odds_format
        }
        
        try:
            response = requests.get(url, params=params, headers=self.headers)
            response.raise_for_status()
            return response.json()
            
        except requests.exceptions.RequestException as e:
            print(f"Error fetching props for event {event_id}: {e}")
            return None
    
    def process_player_data(self, player_name: str, category: str, prop_data: Dict):
        """
        Process and store unique player data with their props
        """
        if player_name not in self.unique_players:
            self.unique_players[player_name] = {
                'categories': set(),
                'props': [],
                'games': set(),
                'bookmakers': set()
            }
        
        # Add category
        self.unique_players[player_name]['categories'].add(category)
        
        # Add prop details
        prop_info = {
            'category': category,
            'bet_type': prop_data['Bet_Type'],
            'line': prop_data['Line'],
            'odds': prop_data['Odds'],
            'bookmaker': prop_data['Bookmaker'],
            'game': prop_data['Game']
        }
        self.unique_players[player_name]['props'].append(prop_info)
        self.unique_players[player_name]['games'].add(prop_data['Game'])
        self.unique_players[player_name]['bookmakers'].add(prop_data['Bookmaker'])
    
    def fetch_all_props_to_excel(self, max_events: Optional[int] = None, 
                                delay_between_requests: float = 1.5,
                                filename: str = None) -> str:
        """
        Fetch all prop listings and save to Excel with organized sheets
        Modified for GitHub Actions with longer delays and error handling
        """
        print("🚀 Starting to fetch MLB prop data for GitHub Actions...")
        print(f"🎯 Target prop categories: {', '.join(self.prop_categories)}")
        
        # Reset data structures
        self.unique_players = {}
        self.over_props = {category: [] for category in self.prop_categories}
        self.under_props = {category: [] for category in self.prop_categories}
        
        # Get all MLB events
        events = self.get_mlb_events()
        if not events:
            print("❌ No events found")
            return ""
        
        # Limit events if specified (for GitHub Actions)
        if max_events:
            events = events[:max_events]
            print(f"📊 Processing first {len(events)} events")
        
        # Store all data organized by category and bet type
        events_summary = []
        
        for i, event in enumerate(events, 1):
            event_id = event['id']
            home_team = event['home_team']
            away_team = event['away_team']
            commence_time = event['commence_time']
            
            print(f"\n[{i}/{len(events)}] Processing: {away_team} @ {home_team}")
            
            # Store event summary
            events_summary.append({
                'Event_ID': event_id,
                'Away_Team': away_team,
                'Home_Team': home_team,
                'Game_Time': commence_time,
                'Matchup': f"{away_team} @ {home_team}"
            })
            
            # Fetch props for this event with retry logic
            max_retries = 3
            props_data = None
            
            for retry in range(max_retries):
                try:
                    props_data = self.get_event_props(event_id, self.prop_categories)
                    if props_data:
                        break
                except Exception as e:
                    print(f"   ⚠️ Retry {retry + 1}/{max_retries} for event {event_id}: {e}")
                    if retry < max_retries - 1:
                        time.sleep(2)  # Wait before retry
                    continue
            
            if props_data:
                available_markets = set()
                
                # Process each bookmaker's data
                for bookmaker in props_data.get('bookmakers', []):
                    bookmaker_name = bookmaker['title']
                    
                    for market in bookmaker.get('markets', []):
                        market_key = market['key']
                        available_markets.add(market_key)
                        
                        # Process each outcome (player prop)
                        for outcome in market.get('outcomes', []):
                            prop_data = {
                                'Event_ID': event_id,
                                'Game': f"{away_team} @ {home_team}",
                                'Game_Time': commence_time,
                                'Category': market_key,
                                'Player': outcome.get('description', 'Unknown'),
                                'Bet_Type': outcome['name'],  # Over/Under
                                'Line': outcome.get('point', 'N/A'),
                                'Odds': outcome['price'],
                                'Bookmaker': bookmaker_name,
                                'Last_Update': market.get('last_update', bookmaker.get('last_update', ''))
                            }
                            
                            # Process player data
                            player_name = prop_data['Player']
                            self.process_player_data(player_name, market_key, prop_data)
                            
                            # Separate Over and Under props
                            if market_key in self.prop_categories:
                                if outcome['name'].lower() == 'over':
                                    self.over_props[market_key].append(prop_data)
                                elif outcome['name'].lower() == 'under':
                                    self.under_props[market_key].append(prop_data)
                
                matching_markets = available_markets.intersection(set(self.prop_categories))
                print(f"   ✅ Found {len(matching_markets)} categories: {', '.join(matching_markets)}")
            else:
                print("   ❌ No props data available")
            
            # Add delay to avoid rate limiting (longer for GitHub Actions)
            if i < len(events):
                time.sleep(delay_between_requests)
        
        # Create Excel file with timestamp
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M")
            filename = f"mlb_props_{timestamp}.xlsx"
        
        try:
            self.create_excel_workbook(events_summary, filename)
            print(f"\n✅ Completed! Processed {len(events)} events")
            print(f"📊 Found {len(self.unique_players)} unique players")
            print(f"📊 Excel file created: {filename}")
            return filename
        except Exception as e:
            print(f"❌ Error creating Excel file: {e}")
            return ""
    
    # ... (rest of the Excel creation methods remain the same as in your original code)
    def create_excel_workbook(self, events_summary: List[Dict], filename: str):
        """
        Create comprehensive Excel workbook with organized sheets
        """
        print(f"📊 Creating Excel workbook: {filename}")
        
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Sheet 1: Summary Dashboard
        self.create_summary_sheet(wb, events_summary)
        
        # Sheet 2: Unique Players List
        self.create_players_sheet(wb)
        
        # Sheet 3: All Over Props Combined
        self.create_over_under_sheet(wb, "All Over", self.over_props)
        
        # Sheet 4: All Under Props Combined
        self.create_over_under_sheet(wb, "All Under", self.under_props)
        
        # Sheet 5: All Props Combined (Over + Under)
        self.create_all_props_combined_sheet(wb)
        
        # Sheet 6: Combined Props Analysis
        self.create_combined_analysis_sheet(wb)
        
        # Sheet 7: Events Summary
        self.create_events_sheet(wb, events_summary)
        
        # Individual sheets for each category - OVER props only (12 sheets)
        for category in self.prop_categories:
            if self.over_props[category]:
                self.create_individual_category_sheet(wb, category, "Over", self.over_props[category])
        
        # Individual sheets for each category - UNDER props only (12 sheets)
        for category in self.prop_categories:
            if self.under_props[category]:
                self.create_individual_category_sheet(wb, category, "Under", self.under_props[category])
        
        # Save workbook
        wb.save(filename)
        print(f"✅ Excel workbook saved: {filename}")
    
    def create_summary_sheet(self, wb, events_summary: List[Dict]):
        """
        Create enhanced summary dashboard sheet
        """
        ws = wb.create_sheet("📊 Enhanced Summary", 0)
        
        # Title
        ws['A1'] = "MLB Props Data - Enhanced Analysis"
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        ws.merge_cells('A1:H1')
        
        # Generation info
        row = 3
        ws[f'A{row}'] = "Report Generated:"
        ws[f'B{row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        row += 1
        
        ws[f'A{row}'] = "Total Events:"
        ws[f'B{row}'] = len(events_summary)
        row += 1
        
        ws[f'A{row}'] = "Total Unique Players:"
        ws[f'B{row}'] = len(self.unique_players)
        row += 2
        
        # Over/Under Summary
        ws[f'A{row}'] = "📋 OVER/UNDER PROPS SUMMARY"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        # Headers for over/under summary
        headers = ['Category', 'Over Props', 'Under Props', 'Total Props', 'Players with Props', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        row += 1
        
        # Category data with over/under breakdown
        for category in self.prop_categories:
            over_count = len(self.over_props[category])
            under_count = len(self.under_props[category])
            total_count = over_count + under_count
            
            # Count unique players for this category
            players_in_category = set()
            for player, data in self.unique_players.items():
                if category in data['categories']:
                    players_in_category.add(player)
            
            if total_count > 0:
                status = "✅ Active"
                color = self.category_colors.get(category, 'FFFFFF')
            else:
                status = "❌ No Data"
                color = 'FFE6E6'
            
            ws.cell(row=row, column=1, value=category).border = self.border
            ws.cell(row=row, column=2, value=over_count).border = self.border
            ws.cell(row=row, column=3, value=under_count).border = self.border
            ws.cell(row=row, column=4, value=total_count).border = self.border
            ws.cell(row=row, column=5, value=len(players_in_category)).border = self.border
            ws.cell(row=row, column=6, value=status).border = self.border
            
            # Color code the row
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            
            row += 1
        
        # Auto-adjust column widths
        self.auto_adjust_columns(ws, max_width=50)
    
    def create_players_sheet(self, wb):
        """
        Create unique players list sheet with their props summary
        """
        ws = wb.create_sheet("👥 Unique Players")
        
        # Title
        ws['A1'] = f"Unique Players List ({len(self.unique_players)} players)"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:G1')
        
        if not self.unique_players:
            ws['A3'] = "No players found"
            return
        
        # Headers
        headers = ['Player Name', 'Categories Count', 'Props Count', 'Games Count', 'Bookmakers', 'Categories', 'Sample Props']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        
        # Player data
        row = 4
        for player_name, player_data in sorted(self.unique_players.items()):
            ws.cell(row=row, column=1, value=player_name).border = self.border
            ws.cell(row=row, column=2, value=len(player_data['categories'])).border = self.border
            ws.cell(row=row, column=3, value=len(player_data['props'])).border = self.border
            ws.cell(row=row, column=4, value=len(player_data['games'])).border = self.border
            ws.cell(row=row, column=5, value=', '.join(player_data['bookmakers'])).border = self.border
            ws.cell(row=row, column=6, value=', '.join(sorted(player_data['categories']))).border = self.border
            
            # Sample props (first 3)
            sample_props = []
            for prop in player_data['props'][:3]:
                sample_props.append(f"{prop['category']} {prop['bet_type']} {prop['line']}")
            sample_text = '; '.join(sample_props)
            if len(player_data['props']) > 3:
                sample_text += f" ... (+{len(player_data['props'])-3} more)"
            
            ws.cell(row=row, column=7, value=sample_text).border = self.border
            row += 1
        
        # Auto-adjust column widths
        self.auto_adjust_columns(ws, max_width=60)
    
    def create_over_under_sheet(self, wb, bet_type: str, props_data: Dict):
        """
        Create separate sheets for Over and Under props
        """
        ws = wb.create_sheet(f"📈 {bet_type} Props")
        
        # Combine all data for this bet type
        all_data = []
        for category, data in props_data.items():
            all_data.extend(data)
        
        if not all_data:
            ws['A1'] = f"No {bet_type.lower()} props data available"
            return
        
        # Title
        ws['A1'] = f"{bet_type} Props Summary ({len(all_data)} props)"
        ws['A1'].font = Font(size=14, bold=True)
        fill_color = self.over_fill if bet_type == "Over" else self.under_fill
        ws['A1'].fill = fill_color
        ws.merge_cells('A1:J1')
        
        # Create DataFrame
        df = pd.DataFrame(all_data)
        df = df.sort_values(['Category', 'Game', 'Player'])
        
        # Write headers
        headers = list(df.columns)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        
        # Write data with color coding
        for row_idx, row_data in enumerate(df.itertuples(index=False), 4):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.border
                cell.font = self.data_font
                
                # Color code by bet type
                if col_idx == 6:  # Bet_Type column
                    cell.fill = fill_color
        
        # Auto-adjust column widths
        self.auto_adjust_columns(ws, max_width=30)
    
    def create_combined_analysis_sheet(self, wb):
        """
        Create combined analysis comparing Over vs Under props
        """
        ws = wb.create_sheet("📊 Over vs Under Analysis")
        
        # Title
        ws['A1'] = "Over vs Under Props Analysis"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:F1')
        
        # Headers
        headers = ['Category', 'Over Count', 'Under Count', 'Total', 'Over %', 'Balance']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        
        row = 4
        for category in self.prop_categories:
            over_count = len(self.over_props[category])
            under_count = len(self.under_props[category])
            total = over_count + under_count
            
            if total > 0:
                over_pct = (over_count / total) * 100
                if over_count > under_count:
                    balance = "Over Heavy"
                elif under_count > over_count:
                    balance = "Under Heavy"
                else:
                    balance = "Balanced"
            else:
                over_pct = 0
                balance = "No Data"
            
            ws.cell(row=row, column=1, value=category).border = self.border
            ws.cell(row=row, column=2, value=over_count).border = self.border
            ws.cell(row=row, column=3, value=under_count).border = self.border
            ws.cell(row=row, column=4, value=total).border = self.border
            ws.cell(row=row, column=5, value=f"{over_pct:.1f}%").border = self.border
            ws.cell(row=row, column=6, value=balance).border = self.border
            
            # Color coding based on balance
            if over_count > under_count:
                fill_color = self.over_fill
            elif under_count > over_count:
                fill_color = self.under_fill
            else:
                fill_color = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
            
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = fill_color
            
            row += 1
        
        # Auto-adjust column widths
        self.auto_adjust_columns(ws, max_width=25)
    
    def create_all_props_combined_sheet(self, wb):
        """
        Create sheet with all props (Over + Under) combined
        """
        ws = wb.create_sheet("📋 All Props Combined")
        
        # Combine all over and under data
        all_data = []
        for category in self.prop_categories:
            all_data.extend(self.over_props[category])
            all_data.extend(self.under_props[category])
        
        if not all_data:
            ws['A1'] = "No props data available"
            return
        
        # Title
        ws['A1'] = f"All Props Combined ({len(all_data)} total props)"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        ws['A1'].font = Font(size=14, bold=True, color='FFFFFF')
        ws.merge_cells('A1:J1')
        
        # Create DataFrame
        df = pd.DataFrame(all_data)
        df = df.sort_values(['Category', 'Player', 'Bet_Type', 'Bookmaker'])
        
        # Write headers
        headers = list(df.columns)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        
        # Write data with color coding
        for row_idx, row_data in enumerate(df.itertuples(index=False), 4):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.border
                cell.font = self.data_font
                
                # Color code by bet type and category
                if col_idx == 6:  # Bet_Type column
                    if str(value).lower() == 'over':
                        cell.fill = self.over_fill
                    elif str(value).lower() == 'under':
                        cell.fill = self.under_fill
                elif col_idx == 4:  # Category column
                    category = str(value)
                    color = self.category_colors.get(category, 'FFFFFF')
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        
        # Auto-adjust column widths
        self.auto_adjust_columns(ws, max_width=30)
    
    def create_individual_category_sheet(self, wb, category: str, bet_type: str, data: List[Dict]):
        """
        Create individual sheet for specific category and bet type (Over or Under only)
        """
        clean_name = category.replace('_', ' ').title()
        sheet_name = f"{bet_type[:1]}-{clean_name}"[:31]  # O-Category or U-Category
        
        ws = wb.create_sheet(sheet_name)
        
        if not data:
            ws['A1'] = f"No {bet_type.lower()} data available for {category}"
            return
        
        # Title with appropriate color
        fill_color = self.over_fill if bet_type == "Over" else self.under_fill
        ws['A1'] = f"{clean_name} - {bet_type} Props ({len(data)} props)"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].fill = fill_color
        ws.merge_cells('A1:J1')
        
        # Create DataFrame
        df = pd.DataFrame(data)
        df = df.sort_values(['Game', 'Player', 'Bookmaker'])
        
        # Write headers
        headers = list(df.columns)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        
        # Write data
        for row_idx, row_data in enumerate(df.itertuples(index=False), 4):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.border
                cell.font = self.data_font
                
                # Color code the bet type column
                if col_idx == 6:  # Bet_Type column
                    cell.fill = fill_color
        
        # Auto-adjust column widths
        self.auto_adjust_columns(ws, max_width=25)
    
    def create_events_sheet(self, wb, events_summary: List[Dict]):
        """
        Create events summary sheet
        """
        ws = wb.create_sheet("🏟️ Events")
        
        if not events_summary:
            ws['A1'] = "No events data available"
            return
        
        # Create DataFrame
        df = pd.DataFrame(events_summary)
        
        # Title
        ws['A1'] = f"MLB Events Summary ({len(events_summary)} games)"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:E1')
        
        # Write headers
        headers = list(df.columns)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        
        # Write data
        for row_idx, row_data in enumerate(df.itertuples(index=False), 4):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.border
                cell.font = self.data_font
        
        # Auto-adjust column widths
        self.auto_adjust_columns(ws, max_width=40)
    
    def get_unique_players_list(self) -> Dict:
        """
        Return the unique players dictionary for further processing
        """
        return self.unique_players
    
    def print_player_summary(self):
        """
        Print a summary of unique players and their props
        """
        print(f"\n{'='*60}")
        print(f"📋 UNIQUE PLAYERS SUMMARY")
        print(f"{'='*60}")
        print(f"Total Unique Players: {len(self.unique_players)}")
        
        for i, (player_name, player_data) in enumerate(sorted(self.unique_players.items()), 1):
            print(f"\n{i:3d}. {player_name}")
            print(f"     Categories: {', '.join(sorted(player_data['categories']))}")
            print(f"     Props Count: {len(player_data['props'])}")
            print(f"     Games: {len(player_data['games'])}")
            print(f"     Bookmakers: {', '.join(player_data['bookmakers'])}")


# Copy all the MLBPropsStatsAnalyzer class code from your original file here
# (I'll include the key parts that need modification for GitHub Actions)

class MLBPropsStatsAnalyzer:
    def __init__(self, excel_file_path: str, max_workers: int = 4):  # Reduced for GitHub Actions
        """
        Initialize the analyzer with the path to the props Excel file
        Modified for GitHub Actions with reduced threading
        """
        self.excel_file_path = excel_file_path
        self.max_workers = max_workers  # Reduced for GitHub Actions environment
        self.props_data = None
        
        # Enhanced caching - stores stats by player + category combination
        self.player_stats_cache = {}
        self.player_id_cache = {}
        self.game_logs_cache = {}
        
        # Thread lock for cache operations
        self.cache_lock = threading.Lock()
        
        # Mapping of prop categories to stat column names
        self.stat_mappings = {
            'batter_home_runs': 'HR',
            'batter_hits': 'H', 
            'batter_total_bases': 'TB',
            'batter_rbis': 'RBI',
            'batter_runs_scored': 'R',
            'batter_hits_runs_rbis': None,
            'batter_singles': None,
            'batter_strikeouts': 'SO',
            'pitcher_strikeouts': 'SO',
            'pitcher_hits_allowed': 'H',
            'pitcher_earned_runs': 'ER',
            'pitcher_outs': None
        }
        
        # Excel styling
        self.header_font = Font(bold=True, color='FFFFFF', size=12)
        self.header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        self.data_font = Font(size=10)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Colors for performance indicators
        self.excellent_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
        self.good_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
        self.average_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
        self.poor_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
        self.bad_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    
    # ... (include all the methods from your original MLBPropsStatsAnalyzer class)
    # For brevity, I'm including just the key modified methods:
    
    def run_full_analysis(self, output_filename: str = None):
        """
        Run the complete analysis pipeline for GitHub Actions
        """
        start_time = time.time()
        
        # Step 1: Load props data
        if not self.load_props_data():
            return None
        
        # Step 2: Process all props with ENHANCED REAL game data
        analyzed_data = self.process_all_props()
        if analyzed_data is None or analyzed_data.empty:
            print("❌ No data to analyze")
            return None
        
        # Step 3: Create enhanced cheat sheets Excel
        if output_filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M")
            output_filename = f"MLB_Props_Cheat_Sheets_{timestamp}.xlsx"
        
        output_file = self.create_cheat_sheets_excel(analyzed_data, output_filename)
        
        end_time = time.time()
        total_time = end_time - start_time
        
        # Print enhanced summary
        print(f"\n{'='*80}")
        print("📊 GITHUB ACTIONS ANALYSIS COMPLETE")
        print(f"{'='*80}")
        print(f"Total Props Analyzed: {len(analyzed_data)}")
        print(f"Unique Players: {analyzed_data['Player'].nunique()}")
        print(f"Categories Covered: {analyzed_data['Category'].nunique()}")
        print(f"Output File: {output_file}")
        print(f"⏱️  Total Processing Time: {total_time/60:.1f} minutes")
        
        return output_file
    
    # Include all other methods from your original class here...
    # (load_props_data, process_all_props, create_cheat_sheets_excel, etc.)


def main():
    """
    Enhanced main function for GitHub Actions
    """
    # Get API key from environment variable (secure for GitHub Actions)
    API_KEY = os.getenv('ODDS_API_KEY')
    
    if not API_KEY:
        print("❌ ODDS_API_KEY environment variable not set")
        print("Please set your API key as a GitHub secret")
        sys.exit(1)
    
    print("🚀 MLB PROPS DATA FETCHER - GITHUB ACTIONS VERSION")
    print("="*70)
    
    # Initialize the fetcher
    fetcher = OddsAPIExcelFetcher(API_KEY)
    
    # Add timestamp to filename for daily runs
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    
    # Fetch all props data and create Excel file
    # Reduced delays and events for GitHub Actions environment
    excel_filename = fetcher.fetch_all_props_to_excel(
        max_events=None,  # Process all events
        delay_between_requests=1.5,  # Slightly longer delay for stability
        filename=f"mlb_props_{timestamp}.xlsx"
    )
    
    if excel_filename:
        print(f"✅ Props data file created: {excel_filename}")
        
        # Get unique players summary
        unique_players = fetcher.get_unique_players_list()
        print(f"📊 Found {len(unique_players)} unique players")
        
        # Run enhanced analysis with reduced threading for GitHub Actions
        print(f"\n🔄 Starting enhanced analysis...")
        analyzer = MLBPropsStatsAnalyzer(excel_filename, max_workers=4)  # Reduced for GitHub Actions
        result_file = analyzer.run_full_analysis(f"MLB_Props_Cheat_Sheets_{timestamp}.xlsx")
        
        if result_file:
            print(f"✅ Analysis complete: {result_file}")
            print(f"📁 Files generated:")
            print(f"   - {excel_filename} (Raw props data)")
            print(f"   - {result_file} (Analysis results)")
        else:
            print("❌ Analysis failed")
            sys.exit(1)
    else:
        print("❌ Failed to create props data file")
        sys.exit(1)


if __name__ == "__main__":
    main()